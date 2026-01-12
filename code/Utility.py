import math
from copy import deepcopy
from enums import EquipLen
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from data import HeaterParam, HeatExchangerParam, CompressorParam, ReactParam
from openpyxl import load_workbook


def checkType(name): 
	length = len(name)
	
	for i in range(0, length):
		# Mixer part
		if (length - i >= EquipLen.MIXER and name[i:i + EquipLen.MIXER] == "MIXER"):
			return "MIX"
		elif (length - i >= EquipLen.SSPLIT and name[i:i + EquipLen.SSPLIT] == "SSPLIT"):
			return "MIX"
		elif (length - i >= EquipLen.FSPLIT and name[i:i + EquipLen.FSPLIT] == "FSPLIT"):
			return "MIX"
		# Seperate part
		elif (length - i >= EquipLen.FLASH2 and name[i:i + EquipLen.FLASH2] == "FLASH2"):
			return "FLASH"
		elif (length - i >= EquipLen.FLASH3 and name[i:i + EquipLen.FLASH3] == "FLASH3"):
			return "FLASH"
		elif (length - i >= EquipLen.SEP and name[i:i + EquipLen.SEP] == "SEP"):
			return "SEP"
		elif (length - i >= EquipLen.SEP2 and name[i:i + EquipLen.SEP2] == "SEP2"):
			return "SEP"
		# Exchanger part
		elif (length - i >= EquipLen.HEATER and name[i:i + EquipLen.HEATER] == "HEATER"):
			return "HTX"
		elif (length - i >= EquipLen.HEATX and name[i:i + EquipLen.HEATX] == "HEATX"):
			return "HEX"
		elif (length - i >= EquipLen.MHEATX and name[i:i + EquipLen.MHEATX] == "MHEATX"):
			return "HEX"
		elif (length - i >= EquipLen.HXFLUX and name[i:i + EquipLen.HXFLUX] == "HXFLUX"):
			return "HEX"
		# Column part
		elif (length - i >= EquipLen.DSTWU and name[i:i + EquipLen.DSTWU] == "DSTWU"):
			return "DIST"
		elif (length - i >= EquipLen.DISTL and name[i:i + EquipLen.DISTL] == "DISTL"):
			return "DIST"
		elif (length - i >= EquipLen.RADFRAC and name[i:i + EquipLen.RADFRAC] == "RADFRAC"):
			return "DIST"
		elif (length - i >= EquipLen.PETROFRAC and name[i:i + EquipLen.PETROFRAC] == "PETROFRAC"):
			return "DIST"
		# Reactor Part
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "RSTOIC"):
			return "REACT"
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "RYIELD"):
			return "REACT"
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "REQUIL"):
			return "REACT"
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "RGIBBS"):
			return "REACT"
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "RCSTR"):
			return "REACT"
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "RPLUG"):
			return "REACT"
		# Pressure Changer Part
		elif (length - i >= EquipLen.COMPR and name[i:i + EquipLen.COMPR] == "COMPR"):
			return "COMP"
		elif (length - i >= EquipLen.PUMP and name[i:i + EquipLen.PUMP] == "MCOMPR"):
			return "COMP"
		# 아래 두 개는 계산 아스펜에서 해줘서 원래 벨브 처리 어떻게 했는지 확인필요
		elif (length - i >= EquipLen.PUMP and name[i:i + EquipLen.PUMP] == "PUMP"):
			return "VALVE"
		elif (length - i >= EquipLen.PUMP and name[i:i + EquipLen.PUMP] == "VALVE"):
			return "VALVE"
		# Solid Seperator Part -> aspen에서 계산해줌
		elif (length - i >= EquipLen.CYCLONE and name[i:i + EquipLen.CYCLONE] == "CYCLONE"):
			return "DIST"
		elif (length - i >= EquipLen.CYCLONE and name[i:i + EquipLen.CYCLONE] == "HYCYC"):
			return "DIST"
	return "ETC"

def calMaterialWeight(material):
	atom = ""
	count = 0
	weight = 0
	atomicWeight = {'H' :1.008, 'He' :4.0026, 'Li' :6.94, 'Be' :9.0122, 'B' :10.81, 'C' :12.011, 'N' :14.007, 'O' :15.999, 'F' :18.998, 'Ne' :20.18, 'Na' :22.99, 'Mg' :24.305, 'Al' :26.982, 'Si' :28.085, 'P' :30.974, 'S' :32.06, 'Cl' :35.45, 'Ar' :39.948, 'K' :39.098, 'Ca' :40.078, 'Sc' :44.956, 'Ti' :47.867, 'V' :50.942, 'Cr' :51.996, 'Mn' :54.938, 'Fe' :55.845, 'Co' :58.933, 'Ni' :58.693, 'Cu' :63.546, 'Zn' :65.38, 'Ga' :69.723, 'Ge' :72.63, 'As' :74.922, 'Se' :78.971, 'Br' :79.904, 'Kr' :83.798, 'Rb' :85.468, 'Sr' :87.62, 'Y' :88.906, 'Zr' :91.224, 'Nb' :92.906, 'Mo' :95.95, 'Tc' :98, 'Ru' :101.07, 'Rh' :102.91, 'Pd' :106.42, 'Ag' :107.87, 'Cd' :112.41, 'In' :114.82, 'Sn' :118.71, 'Sb' :121.76, 'Te' :127.6, 'I' :126.9, 'Xe' :131.29, 'Cs' :132.91, 'Ba' :137.33, 'La' :138.91, 'Ce' :140.12, 'Pr' :140.91, 'Nd' :144.24, 'Pm' :145, 'Sm' :150.36, 'Eu' :151.96, 'Gd' :157.25, 'Tb' :158.93, 'Dy' :162.5, 'Ho' :164.93, 'Er' :167.26, 'Tm' :168.93, 'Yb' :173.05, 'Lu' :174.97, 'Hf' :178.49, 'Ta' :180.95, 'W' :183.84, 'Re' :186.21, 'Os' :190.23, 'Ir' :192.22, 'Pt' :195.08, 'Au' :196.97, 'Hg' :200.59, 'Tl' :204.38, 'Pb' :207.2, 'Bi' :208.98, 'Po' :209, 'At' :210, 'Rn' :222, 'Fr' :223, 'Ra' :226, 'Ac' :227, 'Th' :232.04, 'Pa' :231.04, 'U' :238.03, 'Np' :237, 'Pu' :244, 'Am' :243, 'Cm' :247, 'Bk' :247, 'Cf' :251, 'Es' :252, 'Fm' :257}
	for i in range(len(material)):
		if material[i].isalpha() and material[i].isupper():
			if atom == "":
				atom = material[i]
			else:
				if (count == 0):
					count += 1
				weight += atomicWeight[atom] * count
				count = 0
				atom = material[i]
		elif material[i].isalpha() and material[i].islower():
			atom += material[i]
		elif material[i].isdigit():
			count = count * 10 + int(material[i])
	if (atom != ""):
		if (count == 0):
			count += 1
		weight += atomicWeight[atom] * count
	return (weight)
	

# cost = {} # 2차원 딕셔너리로 "이름" : {딕셔너리} 이렇게 저장하고 각 유닛 종류별 인자와 계산 결과를 출력한다.
def calEquipmentCost(inputData, cost, utility): #react도 추가해야함.
	for key in inputData:
		temp = {}
		type = inputData[key]["Type"]
 		# 여기서 이제 cost의 값들을 하나씩 이름, 인자 순으로 저장해야함.
		'''
		1. EQUIPMENT COST
		- HEX, HTX : ((10^(K1+K2+K3))*(Capacity / 10)^(0.6)) * (CEPCI(June 2024) / CEPCI(Sept 2001))
		- COMP : (10^(K1 + K2 * log(Capacity) + K3 * ((log(Capacity))^2))) * (CEPCI(June 2024) / CEPCI(Sept 2001))
		2. C_BM 
		- HEX, HTX : EQUIPMENT COST * (B1 + B2*FM)
		'''
		if (type == "HTX"):
			if ("HOT UTILITY[kW]" in utility[key]):
				utilityKey = "HOT UTILITY[kW]"
			elif ("ELECTRICITY UTILITY[kW]" in utility[key]):
				utilityKey = "ELECTRICITY UTILITY[kW]"
			else:
				utilityKey = "COOLING UTILITY[kg/hr]"
			# print(key, utility[key][utilityKey])
			temp["Diphenyl heater"] = deepcopy(HeaterParam["Diphenyl heater"])
			temp["Diphenyl heater"]["EQUIPMENT COST"] = ((10**(temp["Diphenyl heater"]["K1"]+temp["Diphenyl heater"]["K2"]+temp["Diphenyl heater"]["K3"]))*(utility[key][utilityKey] / 10)**(0.6)) * (798.8 / 397)
			# print(temp["Diphenyl heater"]["EQUIPMENT COST"])
			temp["Molten salt heater"] = deepcopy(HeaterParam["Molten salt heater"])
			temp["Molten salt heater"]["EQUIPMENT COST"] = ((10**(temp["Molten salt heater"]["K1"]+temp["Molten salt heater"]["K2"]+temp["Molten salt heater"]["K3"]))*(utility[key][utilityKey] / 10)**(0.6)) * (798.8 / 397)
			
			temp["Hot water heater"] = deepcopy(HeaterParam["Hot water heater"])
			temp["Hot water heater"]["EQUIPMENT COST"] = ((10**(temp["Hot water heater"]["K1"]+temp["Hot water heater"]["K2"]+temp["Hot water heater"]["K3"]))*(utility[key][utilityKey] / 10)**(0.6)) * (798.8 / 397)
			
			temp["Steam boiler"] = deepcopy(HeaterParam["Steam boiler"])
			temp["Steam boiler"]["EQUIPMENT COST"] = ((10**(temp["Steam boiler"]["K1"]+temp["Steam boiler"]["K2"]+temp["Steam boiler"]["K3"]))*(utility[key][utilityKey] / 10)**(0.6)) * (798.8 / 397)
		elif (type == "HEX"):
			temp["Fixed tube"] = deepcopy(HeatExchangerParam["Fixed tube"])
			temp["Fixed tube"]["EQUIPMENT COST"] = ((10**(temp["Fixed tube"]["K1"]+temp["Fixed tube"]["K2"]+temp["Fixed tube"]["K3"]))*(inputData[key]["HeatTransferArea"] / 10)**(0.6)) * (798.8 / 397)
			
			temp["Floating head"] = deepcopy(HeatExchangerParam["Floating head"])
			temp["Floating head"]["EQUIPMENT COST"] = ((10**(temp["Floating head"]["K1"]+temp["Floating head"]["K2"]+temp["Floating head"]["K3"]))*(inputData[key]["HeatTransferArea"] / 10)**(0.6)) * (798.8 / 397)
			
			temp["U-tube"] = deepcopy(HeatExchangerParam["U-tube"])
			temp["U-tube"]["EQUIPMENT COST"] = ((10**(temp["U-tube"]["K1"]+temp["U-tube"]["K2"]+temp["U-tube"]["K3"]))*(inputData[key]["HeatTransferArea"] / 10)**(0.6)) * (798.8 / 397)
			
			temp["Bayonet"] = deepcopy(HeatExchangerParam["Bayonet"])
			temp["Bayonet"]["EQUIPMENT COST"] = ((10**(temp["Bayonet"]["K1"]+temp["Bayonet"]["K2"]+temp["Bayonet"]["K3"]))*(inputData[key]["HeatTransferArea"] / 10)**(0.6)) * (798.8 / 397)
		elif (type == "COMP"):
			temp["Centrifugal, axial and reciprocating"] = deepcopy(CompressorParam["Centrifugal, axial and reciprocating"])
			if (inputData[key]["DriverPower"] == 0):
				continue
				#error
				# 나중에 여기에 에러 처리 코드 넣기
			temp["Centrifugal, axial and reciprocating"]["EQUIPMENT COST"] = (10**(temp["Centrifugal, axial and reciprocating"]["K1"] + temp["Centrifugal, axial and reciprocating"]["K2"] * (math.log(inputData[key]["DriverPower"], 10)) + (temp["Centrifugal, axial and reciprocating"]["K3"] * ((math.log(inputData[key]["DriverPower"], 10))**2)))) * (798.8 / 397)
			temp["Rotary"] = deepcopy(CompressorParam["Rotary"])
			temp["Rotary"]["EQUIPMENT COST"] = (10**(temp["Rotary"]["K1"] + temp["Rotary"]["K2"] * (math.log(inputData[key]["DriverPower"], 10)) + (temp["Rotary"]["K3"] * ((math.log(inputData[key]["DriverPower"], 10))**2)))) * (798.8 / 397)
		elif (type == "REACT"):
			temp["input"] = deepcopy(ReactParam["Nan"])
			temp["input"]["EQUIPMENT COST"] = 0
		# 여기는 이미 가격 계산 되어있으면 계산 안 하는 부분
		if (inputData[key]["EquipmentCost"] != 0): 
			temp["ATEA"] = {}
			temp["ATEA"]["EQUIPMENT COST"] = inputData[key]["EquipmentCost"]
		cost[key] = deepcopy(temp)
		# print(cost[key])
	 
# 이제 여기서 Capacity 값은 각 모듈별로 파싱해서 저장해둬야함.
def safe_df(df):
    # 인덱스가 RangeIndex(0,1,2,...)가 아니거나, 인덱스 이름이 있다면 컬럼으로 복구
    if not isinstance(df.index, pd.RangeIndex) or df.index.name is not None:
        return df.reset_index()
    return df

def write_block(ws, df, title):
    ws.append([title])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="004085")
    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)
    ws.append([])

def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

def printout(inputData, cost, utility, CAPEX, OPEX, profitAnalysis):
    
    columns = ["Name","EquipmentCost","InstalledCost","EquipmentWeight",
           "InstalledWeight","UtilityCost","HeatTransferArea","DriverPower"]
    wb = Workbook()
    ws = wb.active
    ws.title = "parse"
    # # 데이터 추가
    ws.append(columns)

# 2️⃣ 데이터 추가
    for name, row in inputData.items():
        ws.append([name] + [row[col] for col in columns[1:]])  # 첫 열은 name, 나머지는 columns 순서대로
    autofit(ws)

    # 시트 2: UTILITY
    ws2 = wb.create_sheet("UTILITY")
    for r in dataframe_to_rows(pd.DataFrame(utility), header=True): ws2.append(r)
    autofit(ws2)
    
    # 시트 3: CAPCOST (블럭 스타일)
    ws3 = wb.create_sheet("Equipment Cost")
    for k,v in cost.items():
        # print(k, v)
        df = pd.DataFrame(v).T   # 행열 전치 추가!
        write_block(ws3, df, k)
    autofit(ws3)

    ws4 = wb.create_sheet("CAPEX")

    for key, values in CAPEX.items():
        ws4.append([key] + values)

    autofit(ws4)
    
    # 시트 5: OPEX
    ws5 = wb.create_sheet("OPEX")
    for key, values in OPEX.items():
        ws5.append([key] + values)
    autofit(ws5)
    
    ws6 = wb.create_sheet("Profitability Analysis")
    opex_df = pd.DataFrame([profitAnalysis]).T  if isinstance(profitAnalysis, dict) else pd.DataFrame(profitAnalysis).T 
    for r in dataframe_to_rows(opex_df, header=False):
        ws6.append(r)
    autofit(ws6)
    
    # 마지막에 한 번만 저장
    wb.save("output.xlsx")

def parseReact(excelReactorData):
	filename = "./input/Material/MaterialData.xlsx"
	df = pd.read_excel(io = filename, sheet_name='Reactor', header=1, engine='openpyxl')
	length = len(df)
	for i in range(length):
		name = df.iat[i, 1]
		cost = df.iat[i, 2]
		excelReactorData[name] = cost

def inputReactorName(inputData, cost):
	filename = "./input/Material/MaterialData.xlsx"

	# 기존 엑셀 로드
	wb = load_workbook(filename)
	ws = wb["Reactor"]

	# pandas로 DF 로드 (기존 데이터 확인용)
	df = pd.read_excel(filename, sheet_name='Reactor', header=1, engine='openpyxl')

	# name 컬럼이 실제 몇 번째인지 찾기
	name_col_index = list(df.columns).index("Reactor Name") + 1  # excel column index (+1 필요)

	current_row = 3  # 데이터 시작 row

	for key in cost:
		if inputData[key]["Type"] == "REACT":
			ws.cell(row=current_row, column=name_col_index).value = key
			current_row += 1

	wb.save(filename)
			

def inputRTX(inputData, cost):
	excelReactorData = {};
	parseReact(excelReactorData)
	for key in cost:
		if (inputData[key]["Type"] == "REACT"):
			if key in excelReactorData:
				cost[key]["input"]["EQUIPMENT COST"] = excelReactorData[key]
			else:
				inputReactorName(inputData, cost)
				raise Exception("reactor의 가격을 엑셀 시트에 이름에 맞게 입력해주세요")
